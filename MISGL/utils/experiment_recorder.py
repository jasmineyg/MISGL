# coding=utf-8

"""Append cross-validation experiment summaries to an Excel workbook."""

import datetime
import os


DEFAULT_EXPERIMENT_RESULT_EXCEL_PATH = '/data/yg/Subgraph-MIL/diffpool2/experiment_results.xlsx'

_METRIC_ORDER = (
  ('acc', 'acc'),
  ('F1', 'f1'),
  ('rec', 'recall'),
  ('prec', 'precision'),
)

_BASE_COLUMNS = [
  'data_name',
  'timestamp',
  'logged_at',
  'cv_num_folds',
  'cv_seed',
  'test_acc_mean',
  'test_acc_std',
]

_SPLIT_COLUMNS = [
  f'{split}_{metric_name}_{stat}'
  for split in ('train', 'val', 'test')
  for _, metric_name in _METRIC_ORDER
  for stat in ('mean', 'std')
]

_TAIL_COLUMNS = [
  'split_path',
  'result_path',
]

_COLUMNS = _BASE_COLUMNS + [
  column for column in _SPLIT_COLUMNS if column not in _BASE_COLUMNS
] + _TAIL_COLUMNS

_LEADERBOARD_COLUMNS = [
  'rank',
  'data_name',
  'timestamp',
  'logged_at',
  'test_acc_mean',
  'test_acc_std',
  'test_f1_mean',
  'test_recall_mean',
  'test_precision_mean',
  'val_acc_mean',
  'train_acc_mean',
  'result_path',
]


def _metric_value(split_summary, split_name, metric_key, stat):
  split_metrics = split_summary.get(split_name, {}) or {}
  metric_stats = split_metrics.get(metric_key, {}) or {}
  value = metric_stats.get(stat, None)
  return None if value is None else float(value)


def _build_row(hparams, data_name, split_summary, split_path=None, result_path=None):
  row = {
    'data_name': data_name,
    'timestamp': str(getattr(hparams, 'timestamp', '') or ''),
    'logged_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'cv_num_folds': int(getattr(hparams, 'cv_num_folds', 10)),
    'cv_seed': int(getattr(hparams, 'cv_seed', 1024)),
    'split_path': split_path,
    'result_path': result_path,
  }

  for split_name in ('train', 'val', 'test'):
    for metric_key, metric_name in _METRIC_ORDER:
      for stat in ('mean', 'std'):
        row[f'{split_name}_{metric_name}_{stat}'] = _metric_value(
          split_summary, split_name, metric_key, stat
        )

  row['test_acc_mean'] = _metric_value(split_summary, 'test', 'acc', 'mean')
  row['test_acc_std'] = _metric_value(split_summary, 'test', 'acc', 'std')
  return row


def _ensure_headers(ws):
  if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
    ws.append(_COLUMNS)
    return list(_COLUMNS)

  headers = [ws.cell(1, col_idx).value for col_idx in range(1, ws.max_column + 1)]
  headers = [str(header) for header in headers if header is not None]
  for column in _COLUMNS:
    if column not in headers:
      headers.append(column)
      ws.cell(1, len(headers)).value = column
  return headers


def _style_sheet(ws, headers):
  from openpyxl.styles import Alignment, Font, PatternFill

  header_fill = PatternFill('solid', fgColor='1F2937')
  header_font = Font(color='FFFFFF', bold=True)
  test_acc_fill = PatternFill('solid', fgColor='C6EFCE')
  rank_fill = PatternFill('solid', fgColor='FFF2CC')

  for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(1, col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

    width = max(12, min(28, len(str(header)) + 2))
    if header in ('data_name', 'timestamp', 'logged_at'):
      width = 22
    elif header in ('split_path', 'result_path'):
      width = 50
    ws.column_dimensions[cell.column_letter].width = width

    if header == 'test_acc_mean':
      ws.column_dimensions[cell.column_letter].width = 16
      for row_idx in range(2, ws.max_row + 1):
        metric_cell = ws.cell(row_idx, col_idx)
        metric_cell.fill = test_acc_fill
        metric_cell.font = Font(bold=True)

    if header == 'rank':
      ws.column_dimensions[cell.column_letter].width = 8
      for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, col_idx).fill = rank_fill

    if header.endswith('_mean') or header.endswith('_std'):
      for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, col_idx).number_format = '0.00%'

  ws.freeze_panes = 'A2'
  ws.auto_filter.ref = ws.dimensions


def _rows_from_sheet(ws, headers):
  rows = []
  for row_idx in range(2, ws.max_row + 1):
    item = {}
    has_value = False
    for col_idx, header in enumerate(headers, start=1):
      value = ws.cell(row_idx, col_idx).value
      item[header] = value
      has_value = has_value or value is not None
    if has_value:
      rows.append(item)
  return rows


def _test_acc_sort_key(row):
  value = row.get('test_acc_mean')
  if value is None:
    return -1.0
  try:
    return float(value)
  except (TypeError, ValueError):
    return -1.0


def _refresh_leaderboard(wb, source_ws, source_headers):
  if 'Leaderboard' in wb.sheetnames:
    ws = wb['Leaderboard']
    ws.delete_rows(1, ws.max_row)
  else:
    ws = wb.create_sheet('Leaderboard', 0)

  ws.append(_LEADERBOARD_COLUMNS)
  rows = sorted(_rows_from_sheet(source_ws, source_headers), key=_test_acc_sort_key, reverse=True)
  for rank, row in enumerate(rows, start=1):
    view_row = dict(row)
    view_row['rank'] = rank
    ws.append([view_row.get(header) for header in _LEADERBOARD_COLUMNS])
  _style_sheet(ws, _LEADERBOARD_COLUMNS)


def append_cv_result(hparams, data_name, split_summary, split_path=None, result_path=None):
  """Create or append one fixed-CV summary row to the configured Excel file."""
  from openpyxl import Workbook, load_workbook

  output_path = getattr(
    hparams, 'experiment_result_excel_path', DEFAULT_EXPERIMENT_RESULT_EXCEL_PATH
  )
  output_path = str(output_path or '').strip()
  if not output_path:
    return None

  output_dir = os.path.dirname(output_path)
  if output_dir:
    os.makedirs(output_dir, exist_ok=True)

  if os.path.exists(output_path):
    wb = load_workbook(output_path)
    ws = wb['CV Results'] if 'CV Results' in wb.sheetnames else wb.active
    ws.title = 'CV Results'
  else:
    wb = Workbook()
    ws = wb.active
    ws.title = 'CV Results'

  headers = _ensure_headers(ws)
  row = _build_row(hparams, data_name, split_summary, split_path=split_path, result_path=result_path)
  ws.append([row.get(header) for header in headers])
  _style_sheet(ws, headers)
  _refresh_leaderboard(wb, ws, headers)
  wb.save(output_path)
  return output_path
