import os
import time
import random
import math
import numpy as np
import pandas as pd
import torch

from MISGL.utils.global_variables import g_key


def _export_subgraph_distance_heatmap_xlsx(D, labels, out_xlsx_path):
  from openpyxl import Workbook
  from openpyxl.formatting.rule import ColorScaleRule
  from openpyxl.utils import get_column_letter
  from openpyxl.styles import Alignment, Font

  M = int(D.shape[0])
  wb = Workbook()
  ws = wb.active
  ws.title = 'Heatmap'
  ws.freeze_panes = 'B2'

  header_font = Font(bold=True)
  center = Alignment(horizontal='center', vertical='center', wrap_text=False)

  ws.cell(row=1, column=1, value=None)

  for j, lab in enumerate(labels, start=2):
    c = ws.cell(row=1, column=j, value=str(lab))
    c.font = header_font
    c.alignment = center

  for i, lab in enumerate(labels, start=2):
    c = ws.cell(row=i, column=1, value=str(lab))
    c.font = header_font
    c.alignment = center

  number_format = '0.000000'
  for i in range(M):
    for j in range(M):
      cell = ws.cell(row=2 + i, column=2 + j, value=float(D[i, j]))
      cell.number_format = number_format
      cell.alignment = center

  if M > 0:
    start_cell = 'B2'
    end_cell = f'{get_column_letter(1 + M)}{1 + M}'
    data_range = f'{start_cell}:{end_cell}'
    dmin = float(np.min(D))
    dmax = float(np.max(D))
    dmid = 0.5 * (dmin + dmax)
    rule = ColorScaleRule(
      start_type='min',
      start_color='0000FF',
      mid_type='num',
      mid_value=dmid,
      mid_color='FFFFFF',
      end_type='max',
      end_color='FF0000',
    )
    ws.conditional_formatting.add(data_range, rule)

  ws.column_dimensions['A'].width = 18
  for col in range(2, 2 + M):
    ws.column_dimensions[get_column_letter(col)].width = 12

  os.makedirs(os.path.dirname(out_xlsx_path), exist_ok=True)
  wb.save(out_xlsx_path)


def export_subgraph_distance_heatmap_from_test_embeddings(
    subgraph_items,
    out_dir,
    filename_prefix,
    sample_frac=0.2,
    seed=1234,
    eps=1e-12,
    out_filename=None,
):
  N = len(subgraph_items)
  if N <= 0:
    raise RuntimeError('未提供任何 test set 子图，无法生成距离热力图。')

  M = int(math.ceil(float(sample_frac) * float(N)))
  M = max(1, min(N, M))
  rng = random.Random(int(seed))
  sampled = [subgraph_items[i] for i in rng.sample(list(range(N)), M)]

  sampled.sort(key=lambda r: (-int(r['correct']), int(r['y_true']), int(r['subgraph_id'])))

  labels = [f"{int(r['subgraph_id'])}_{int(r['correct'])}_{int(r['y_true'])}" for r in sampled]
  G = np.stack([r['g'].astype(np.float32) for r in sampled], axis=0)
  norms = np.linalg.norm(G, axis=1, keepdims=True)
  G_hat = G / (norms + float(eps))
  S = np.matmul(G_hat, G_hat.T)
  D = 1.0 - S
  np.fill_diagonal(D, 0.0)

  os.makedirs(out_dir, exist_ok=True)
  out_name = str(out_filename) if out_filename is not None else f'heatmap_{filename_prefix}.xlsx'
  if not out_name.lower().endswith('.xlsx'):
    out_name = out_name + '.xlsx'
  out_xlsx_path = os.path.join(out_dir, out_name)
  _export_subgraph_distance_heatmap_xlsx(D, labels, out_xlsx_path)
  return out_xlsx_path, int(M)


def _vector_l2_norm(x):
  return torch.sqrt(torch.sum(x * x, dim=-1))


def _vector_dim_var(x):
  mu = torch.mean(x, dim=-1, keepdim=True)
  return torch.mean((x - mu) ** 2, dim=-1)


def _cosine(a, b, eps=1e-12):
  a_norm = _vector_l2_norm(a)
  b_norm = _vector_l2_norm(b)
  denom = a_norm * b_norm + eps
  return torch.sum(a * b, dim=-1) / denom


def _sanitize_sheet_name(name):
  s = str(name)
  for ch in [':', '\\', '/', '?', '*', '[', ']']:
    s = s.replace(ch, '_')
  s = s.strip()
  return s if len(s) > 0 else 'sheet'


def _unique_sheet_name(base_name, used_names):
  base = _sanitize_sheet_name(base_name)
  base = base[:31]
  if base not in used_names:
    used_names.add(base)
    return base
  k = 1
  while True:
    suffix = f'_{k}'
    cut = 31 - len(suffix)
    cand = (base[:cut] + suffix)[:31]
    if cand not in used_names:
      used_names.add(cand)
      return cand
    k += 1


def _map_subgraph_nodes_to_labels(subgraph, node_binary_labels, n_i):
  nodes = list(subgraph.nodes())
  labels = np.zeros(len(nodes), dtype=np.int64)
  total = len(node_binary_labels)
  for j, node_id in enumerate(nodes):
    idx = None
    if isinstance(node_id, (int, np.integer)):
      idx = int(node_id)
    else:
      attr = subgraph.nodes[node_id]
      for key in ('original_index', 'orig_id', 'node_index', 'original_id'):
        if key in attr and attr[key] is not None:
          try:
            idx = int(attr[key])
            break
          except Exception:
            pass
    if idx is not None and 0 <= idx < total:
      labels[j] = int(node_binary_labels[idx])
    else:
      labels[j] = 0
  return labels[:n_i]


def _map_subgraph_nodes_to_orig_ids(subgraph, n_i):
  nodes = list(subgraph.nodes())
  out = np.full(len(nodes), -1, dtype=np.int64)
  for j, node_id in enumerate(nodes):
    attr = subgraph.nodes[node_id]
    idx = None
    for key in ('original_id', 'original_index', 'orig_id', 'node_index'):
      if key in attr and attr[key] is not None:
        try:
          idx = int(attr[key])
          break
        except Exception:
          pass
    if idx is None and isinstance(node_id, (int, np.integer)):
      idx = int(node_id)
    out[j] = -1 if idx is None else idx
  return out[:n_i]


def _get_node_labels_for_batch(dataset_raw, orig_graph_indices, num_list):
  if dataset_raw is None:
    return [np.zeros(n, dtype=np.int64) for n in num_list]
  node_labels_collection = dataset_raw.get('node_binary_labels', None)
  subgraphs = dataset_raw.get('subgraph_structures', None)
  if node_labels_collection is None or subgraphs is None:
    return [np.zeros(n, dtype=np.int64) for n in num_list]

  out = []
  for orig_idx, n_i in zip(orig_graph_indices, num_list):
    if 0 <= orig_idx < len(subgraphs):
      subgraph = subgraphs[orig_idx]
      out.append(_map_subgraph_nodes_to_labels(subgraph, node_labels_collection, n_i))
    else:
      out.append(np.zeros(n_i, dtype=np.int64))
  return out


def _get_node_orig_ids_for_batch(dataset_raw, orig_graph_indices, num_list):
  if dataset_raw is None:
    return [np.full(n, -1, dtype=np.int64) for n in num_list]
  subgraphs = dataset_raw.get('subgraph_structures', None)
  if subgraphs is None:
    return [np.full(n, -1, dtype=np.int64) for n in num_list]

  out = []
  for orig_idx, n_i in zip(orig_graph_indices, num_list):
    if 0 <= orig_idx < len(subgraphs):
      subgraph = subgraphs[orig_idx]
      out.append(_map_subgraph_nodes_to_orig_ids(subgraph, n_i))
    else:
      out.append(np.full(n_i, -1, dtype=np.int64))
  return out


def export_test_result_analyze_excel(
    model,
    test_loader,
    hparams,
    dataset_raw=None,
    output_dir='/data/yg/Subgraph-MIL/diffpool2/results',
    seed=1024,
    sample_frac=0.2,
    filename_prefix=None,
    heatmap_filename=None,
):
  if not hasattr(model, 'forward_with_embeddings'):
    raise RuntimeError('模型未实现 forward_with_embeddings，无法导出最终层 embedding。')

  device = torch.device(getattr(hparams, 'device', 'cpu'))
  model.eval()

  records = []
  heatmap_items = []
  pos_sum = None
  pos_cnt = 0

  with torch.no_grad():
    global_idx = 0
    for graph_data in test_loader:
      ypred_out, emb = model.forward_with_embeddings(graph_data)
      if emb is None:
        raise RuntimeError('forward_with_embeddings 未返回 embedding dict。')

      logits = ypred_out['ypred_A'] if isinstance(ypred_out, dict) and 'ypred_A' in ypred_out else ypred_out
      probs = torch.sigmoid(logits.view(-1)).detach().cpu().tolist()
      preds = [1 if float(p) >= 0.5 else 0 for p in probs]
      ys = graph_data[g_key.y].detach().cpu().view(-1).tolist()

      batch_num_nodes = graph_data[g_key.node_num]
      if isinstance(batch_num_nodes, torch.Tensor):
        num_list = [int(n) for n in batch_num_nodes.detach().cpu().tolist()]
      else:
        num_list = [int(n) for n in batch_num_nodes]

      orig_idx_tensor = graph_data.get(g_key.orig_graph_idx, None)
      if orig_idx_tensor is not None and isinstance(orig_idx_tensor, torch.Tensor):
        orig_graph_indices = [int(i) for i in orig_idx_tensor.detach().cpu().tolist()]
      else:
        orig_graph_indices = list(range(len(ys)))

      subgraph_id_tensor = graph_data.get(g_key.subgraph_id, None)
      if subgraph_id_tensor is not None and isinstance(subgraph_id_tensor, torch.Tensor):
        subgraph_ids = [int(i) for i in subgraph_id_tensor.detach().cpu().tolist()]
      else:
        subgraph_ids = [-1 for _ in ys]

      h = emb.get('h', None)
      if h is None:
        h = emb.get('h_B', None) if emb.get('h_B', None) is not None else emb.get('h_A', None)
      if h is None:
        raise RuntimeError('embedding dict 中缺少节点 embedding（key: h / h_A / h_B）。')

      graph_emb_classifier = emb.get('graph_emb_classifier', None)
      if graph_emb_classifier is None:
        graph_emb_classifier = emb.get('graph_emb_H2', None)
      if graph_emb_classifier is None:
        graph_emb_classifier = emb.get('mean_vec', None)
      if graph_emb_classifier is None:
        graph_emb_classifier = emb.get('graph_emb', None)

      h_dim = int(h.size(-1))

      def _match_graph_emb_dim(t):
        if t is None or (not isinstance(t, torch.Tensor)):
          return None
        if t.dim() != 2:
          return None
        if int(t.size(1)) != h_dim:
          return None
        return t

      graph_emb = (
        _match_graph_emb_dim(emb.get('graph_emb', None)) or
        _match_graph_emb_dim(emb.get('graph_emb_H2', None)) or
        _match_graph_emb_dim(emb.get('mean_vec', None)) or
        _match_graph_emb_dim(graph_emb_classifier)
      )
      if graph_emb is None:
        raise RuntimeError('embedding dict 中缺少可与节点 embedding 对齐的 graph embedding。')

      node_labels_per_graph = _get_node_labels_for_batch(dataset_raw, orig_graph_indices, num_list)

      for i, n_i in enumerate(num_list):
        y_true = int(ys[i])
        y_pred = int(preds[i])
        correct = 1 if y_true == y_pred else 0

        g = graph_emb[i].detach().to(device)
        g_classifier = graph_emb_classifier[i].detach().cpu().numpy().astype(np.float32)
        g_norm = float(_vector_l2_norm(g).detach().cpu().item())
        g_var = float(_vector_dim_var(g).detach().cpu().item())

        node_labels = node_labels_per_graph[i]
        num_pos = int(np.sum(node_labels[:n_i] == 1)) if node_labels is not None else 0

        records.append({
          'global_idx': global_idx,
          'subgraph_id': int(subgraph_ids[i]),
          'ground_truth': y_true,
          'prediction': y_pred,
          'correct': correct,
          'num_nodes': int(n_i),
          'num_pos_nodes': num_pos,
          'graph_emb_norm': g_norm,
          'graph_emb_var': g_var,
        })
        heatmap_items.append({
          'subgraph_id': int(subgraph_ids[i]),
          'y_true': y_true,
          'correct': correct,
          'g': g_classifier,
        })

        if n_i > 0 and node_labels is not None:
          h_i = h[i, :n_i, :].detach().to(device)
          if h_i.dim() != 2 or g.dim() != 1 or h_i.size(1) != g.size(0):
            raise RuntimeError(f'节点 embedding 维度与 graph embedding 不一致：h={tuple(h_i.size())}, g={tuple(g.size())}')
          pos_mask = torch.tensor((node_labels[:n_i] == 1), device=device, dtype=torch.bool)
          if torch.any(pos_mask):
            pos_vecs = h_i[pos_mask]
            s = torch.sum(pos_vecs, dim=0)
            if pos_sum is None:
              pos_sum = s.detach().clone()
            else:
              pos_sum += s.detach()
            pos_cnt += int(pos_vecs.size(0))

        global_idx += 1

  if len(records) == 0:
    raise RuntimeError('test_loader 为空，未导出任何子图。')

  D = int(pos_sum.size(0)) if pos_sum is not None else int(h.size(-1))
  if pos_sum is None or pos_cnt <= 0:
    c_pos = torch.zeros(D, device=device, dtype=torch.float32)
  else:
    c_pos = (pos_sum / float(pos_cnt)).to(dtype=torch.float32)

  N = len(records)
  sample_k = int(round(float(sample_frac) * float(N)))
  sample_k = max(1, min(N, sample_k)) if N > 0 and sample_frac > 0 else 0
  rng = random.Random(int(seed))
  sampled_global_indices = set(rng.sample(list(range(N)), sample_k)) if sample_k > 0 else set()

  if filename_prefix is None:
    filename_prefix = time.strftime('%Y%m%d-%H%M%S')
  output_path = os.path.join(output_dir, f'{filename_prefix}_result_analyze.xlsx')
  os.makedirs(os.path.dirname(output_path), exist_ok=True)

  heatmap_dir = output_dir
  heatmap_seed = int(getattr(hparams, 'subgraph_distance_heatmap_seed', 1234))
  heatmap_sample_frac = float(getattr(hparams, 'subgraph_distance_heatmap_sample_frac', sample_frac))
  heatmap_out_filename = heatmap_filename if heatmap_filename is not None else f'{filename_prefix}_heatmap.xlsx'
  heatmap_xlsx_path, heatmap_M = export_subgraph_distance_heatmap_from_test_embeddings(
    subgraph_items=heatmap_items,
    out_dir=heatmap_dir,
    filename_prefix=filename_prefix,
    sample_frac=heatmap_sample_frac,
    seed=heatmap_seed,
    eps=1e-12,
    out_filename=heatmap_out_filename,
  )

  df_summary = pd.DataFrame([
    {k: r[k] for k in ['subgraph_id', 'ground_truth', 'prediction', 'correct', 'num_nodes', 'num_pos_nodes', 'graph_emb_norm', 'graph_emb_var']}
    for r in records
  ])

  writer = pd.ExcelWriter(output_path, engine='openpyxl')
  df_summary.to_excel(writer, sheet_name='Summary', index=False)

  used_sheet_names = {'Summary'}
  with torch.no_grad():
    global_idx = 0
    for graph_data in test_loader:
      ypred_out, emb = model.forward_with_embeddings(graph_data)
      logits = ypred_out['ypred_A'] if isinstance(ypred_out, dict) and 'ypred_A' in ypred_out else ypred_out
      probs = torch.sigmoid(logits.view(-1)).detach().cpu().tolist()
      preds = [1 if float(p) >= 0.5 else 0 for p in probs]
      ys = graph_data[g_key.y].detach().cpu().view(-1).tolist()

      batch_num_nodes = graph_data[g_key.node_num]
      if isinstance(batch_num_nodes, torch.Tensor):
        num_list = [int(n) for n in batch_num_nodes.detach().cpu().tolist()]
      else:
        num_list = [int(n) for n in batch_num_nodes]

      orig_idx_tensor = graph_data.get(g_key.orig_graph_idx, None)
      if orig_idx_tensor is not None and isinstance(orig_idx_tensor, torch.Tensor):
        orig_graph_indices = [int(i) for i in orig_idx_tensor.detach().cpu().tolist()]
      else:
        orig_graph_indices = list(range(len(ys)))

      subgraph_id_tensor = graph_data.get(g_key.subgraph_id, None)
      if subgraph_id_tensor is not None and isinstance(subgraph_id_tensor, torch.Tensor):
        subgraph_ids = [int(i) for i in subgraph_id_tensor.detach().cpu().tolist()]
      else:
        subgraph_ids = [-1 for _ in ys]

      h = emb.get('h', None)
      if h is None:
        h = emb.get('h_B', None) if emb.get('h_B', None) is not None else emb.get('h_A', None)

      graph_emb_classifier = emb.get('graph_emb_classifier', None)
      if graph_emb_classifier is None:
        graph_emb_classifier = emb.get('graph_emb_H2', None)
      if graph_emb_classifier is None:
        graph_emb_classifier = emb.get('mean_vec', None)
      if graph_emb_classifier is None:
        graph_emb_classifier = emb.get('graph_emb', None)

      h_dim = int(h.size(-1))

      def _match_graph_emb_dim(t):
        if t is None or (not isinstance(t, torch.Tensor)):
          return None
        if t.dim() != 2:
          return None
        if int(t.size(1)) != h_dim:
          return None
        return t

      graph_emb = (
        _match_graph_emb_dim(emb.get('graph_emb', None)) or
        _match_graph_emb_dim(emb.get('graph_emb_H2', None)) or
        _match_graph_emb_dim(emb.get('mean_vec', None)) or
        _match_graph_emb_dim(graph_emb_classifier)
      )

      node_labels_per_graph = _get_node_labels_for_batch(dataset_raw, orig_graph_indices, num_list)
      node_orig_ids_per_graph = _get_node_orig_ids_for_batch(dataset_raw, orig_graph_indices, num_list)

      for i, n_i in enumerate(num_list):
        if global_idx in sampled_global_indices and n_i > 0:
          y_true = int(ys[i])
          y_pred = int(preds[i])
          correct = 1 if y_true == y_pred else 0
          suffix = 'correct' if correct == 1 else 'wrong'
          sheet_name = _unique_sheet_name(f'{int(subgraph_ids[i])}_{suffix}', used_sheet_names)

          g = graph_emb[i].detach().to(device)
          h_i = h[i, :n_i, :].detach().to(device)
          if h_i.dim() != 2 or g.dim() != 1 or h_i.size(1) != g.size(0):
            raise RuntimeError(f'节点 embedding 维度与 graph embedding 不一致：h={tuple(h_i.size())}, g={tuple(g.size())}')

          node_labels = node_labels_per_graph[i]
          node_orig_ids = node_orig_ids_per_graph[i]
          node_ids = list(range(int(n_i)))
          node_labels_list = node_labels[:n_i].astype(np.int64).tolist() if node_labels is not None else [0] * int(n_i)
          is_pos = [bool(v == 1) for v in node_labels_list]
          orig_ids_list = node_orig_ids[:n_i].astype(np.int64).tolist() if node_orig_ids is not None else [-1] * int(n_i)

          node_norm = _vector_l2_norm(h_i).detach().cpu().numpy().astype(np.float64)
          node_var = _vector_dim_var(h_i).detach().cpu().numpy().astype(np.float64)
          cos_to_g = _cosine(h_i, g.unsqueeze(0).expand_as(h_i)).detach().cpu().numpy().astype(np.float64)
          cos_to_pos = _cosine(h_i, c_pos.unsqueeze(0).expand_as(h_i)).detach().cpu().numpy().astype(np.float64)

          df_nodes = pd.DataFrame({
            'node_id': node_ids,
            'original_node_id': orig_ids_list,
            'node_label': node_labels_list,
            'is_pos_node': is_pos,
            'node_emb_norm': node_norm,
            'node_emb_var': node_var,
            'cos_to_graph_emb': cos_to_g,
            'cos_to_pos_center': cos_to_pos,
          })
          df_nodes.to_excel(writer, sheet_name=sheet_name, index=False)

        global_idx += 1

  writer.close()
  return {
    'result_analyze_xlsx': output_path,
    'distance_heatmap_xlsx': heatmap_xlsx_path,
    'distance_heatmap_M': heatmap_M,
  }
