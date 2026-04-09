import os
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union

import numpy as np
import pandas as pd
import networkx as nx


_INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]\:\*\?\/\\]")


try:
  import torch
except Exception:
  torch = None


def _as_cpu_numpy_adj(hatA: Union[Any, np.ndarray]) -> np.ndarray:
  if isinstance(hatA, np.ndarray):
    A = hatA
  elif torch is not None and isinstance(hatA, torch.Tensor):
    t = hatA.detach()
    if t.is_sparse:
      t = t.to_dense()
    if t.dim() == 3 and int(t.size(0)) == 1:
      t = t.squeeze(0)
    A = t.cpu().numpy()
  else:
    raise TypeError(f'Unsupported adjacency type: {type(hatA)}')
  if A.ndim != 2 or A.shape[0] != A.shape[1]:
    raise ValueError(f'Adjacency must be square 2D array, got shape={A.shape}')
  return A.astype(np.float64, copy=False)


def _build_nx_graph_from_adj(
    hatA: Union[Any, np.ndarray],
    node_ids: Optional[List[int]] = None,
    threshold: float = 0.0,
) -> nx.Graph:
  A = _as_cpu_numpy_adj(hatA)
  n = int(A.shape[0])
  if node_ids is None:
    node_ids = list(range(n))
  if len(node_ids) != n:
    raise ValueError(f'node_ids length mismatch: len(node_ids)={len(node_ids)} vs N={n}')

  np.fill_diagonal(A, 0.0)
  rows, cols = np.where(np.triu(A, k=1) > float(threshold))
  G = nx.Graph()
  G.add_nodes_from(node_ids)
  for i, j in zip(rows.tolist(), cols.tolist()):
    G.add_edge(node_ids[int(i)], node_ids[int(j)])
  return G


def _safe_sheet_name(name: str, used: Optional[set] = None) -> str:
  used = used or set()
  raw = str(name) if name is not None else 'graph'
  raw = raw.strip() if raw.strip() else 'graph'
  raw = _INVALID_SHEET_CHARS_RE.sub('_', raw)
  raw = raw[:31]
  if raw == '':
    raw = 'graph'
  if raw not in used:
    used.add(raw)
    return raw
  base = raw[:27] if len(raw) > 27 else raw
  k = 1
  while True:
    cand = f'{base}_{k}'
    cand = cand[:31]
    if cand not in used:
      used.add(cand)
      return cand
    k += 1


def _base_sheet_name(name: str) -> str:
  raw = str(name) if name is not None else 'graph'
  raw = raw.strip() if raw.strip() else 'graph'
  raw = _INVALID_SHEET_CHARS_RE.sub('_', raw)
  raw = raw[:31]
  return raw if raw else 'graph'


def _rank_pearson_corr(x: np.ndarray, y: np.ndarray) -> float:
  if x.size <= 1 or y.size <= 1:
    return float('nan')
  sx = pd.Series(x).rank(method='average').to_numpy(dtype=np.float64)
  sy = pd.Series(y).rank(method='average').to_numpy(dtype=np.float64)
  if float(np.std(sx)) < 1e-12 or float(np.std(sy)) < 1e-12:
    return float('nan')
  return float(np.corrcoef(sx, sy)[0, 1])


def _fmt_list(pairs: List[Tuple[Any, Any]], max_items: int = 10) -> str:
  show = pairs[:max_items]
  return ', '.join([f'({a},{b})' for a, b in show])


def analyze_one_graph(
    hatA_or_G: Union[Any, np.ndarray, nx.Graph],
    graph_name: str,
    node_ids: Optional[List[int]] = None,
    threshold: float = 0.0,
    print_summary: bool = True,
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
  if isinstance(hatA_or_G, nx.Graph):
    G = hatA_or_G
  else:
    G = _build_nx_graph_from_adj(hatA_or_G, node_ids=node_ids, threshold=threshold)

  nodes = list(G.nodes())
  N = int(G.number_of_nodes())
  E = int(G.number_of_edges())

  degree_dict = dict(G.degree())
  degrees = np.array([float(degree_dict[n]) for n in nodes], dtype=np.float64)

  one_hop_counts = np.zeros((N,), dtype=np.int64)
  two_hop_counts = np.zeros((N,), dtype=np.int64)
  for i, v in enumerate(nodes):
    lengths1 = nx.single_source_shortest_path_length(G, v, cutoff=1)
    one_hop_counts[i] = int(len(lengths1))
    lengths = nx.single_source_shortest_path_length(G, v, cutoff=2)
    two_hop_counts[i] = int(len(lengths))
  one_hop_ratios = one_hop_counts.astype(np.float64) / float(max(N, 1))
  two_hop_ratios = two_hop_counts.astype(np.float64) / float(max(N, 1))

  avg_degree = float(np.mean(degrees)) if N > 0 else 0.0
  max_degree = float(np.max(degrees)) if N > 0 else 0.0

  def _stats(arr: np.ndarray) -> Dict[str, float]:
    if arr.size == 0:
      return {'min': float('nan'), 'median': float('nan'), 'mean': float('nan'), 'p90': float('nan'), 'max': float('nan')}
    return {
      'min': float(np.min(arr)),
      'median': float(np.median(arr)),
      'mean': float(np.mean(arr)),
      'p90': float(np.percentile(arr, 90)),
      'max': float(np.max(arr)),
    }

  one_thr_stats = _stats(one_hop_ratios)
  thr_stats = _stats(two_hop_ratios)
  ge_05_prop_1hop = float(np.mean(one_hop_ratios >= 0.5)) if N > 0 else 0.0
  ge_09_prop = float(np.mean(two_hop_ratios >= 0.9)) if N > 0 else 0.0
  corr = _rank_pearson_corr(degrees, two_hop_ratios)

  qs = np.array([0, 25, 50, 75, 90, 95, 99, 100], dtype=np.float64)
  deg_q = np.percentile(degrees, qs).tolist() if N > 0 else [float('nan')] * len(qs)
  thr1_q = np.percentile(one_hop_ratios, qs).tolist() if N > 0 else [float('nan')] * len(qs)
  thr_q = np.percentile(two_hop_ratios, qs).tolist() if N > 0 else [float('nan')] * len(qs)

  order_deg = np.argsort(-degrees) if N > 0 else np.array([], dtype=np.int64)
  top_degree = [(int(nodes[i]), int(degrees[i])) for i in order_deg[:10].tolist()]

  order_thr1_desc = np.argsort(-one_hop_ratios) if N > 0 else np.array([], dtype=np.int64)
  top_thr1 = [(int(nodes[i]), float(one_hop_ratios[i])) for i in order_thr1_desc[:10].tolist()]
  order_thr1_asc = np.argsort(one_hop_ratios) if N > 0 else np.array([], dtype=np.int64)
  bottom_thr1 = [(int(nodes[i]), float(one_hop_ratios[i])) for i in order_thr1_asc[:10].tolist()]
  order_thr_desc = np.argsort(-two_hop_ratios) if N > 0 else np.array([], dtype=np.int64)
  top_thr = [(int(nodes[i]), float(two_hop_ratios[i])) for i in order_thr_desc[:10].tolist()]
  order_thr_asc = np.argsort(two_hop_ratios) if N > 0 else np.array([], dtype=np.int64)
  bottom_thr = [(int(nodes[i]), float(two_hop_ratios[i])) for i in order_thr_asc[:10].tolist()]

  if print_summary:
    print(f'[CoarseGraph] {graph_name}')
    print(f'(1) Graph: N={N}, E={E}, avg_degree={avg_degree:.4f}, max_degree={max_degree:.0f}')
    print(
      '(2) one_hop_ratio: '
      f'min={one_thr_stats["min"]:.4f} / median={one_thr_stats["median"]:.4f} / mean={one_thr_stats["mean"]:.4f} / '
      f'p90={one_thr_stats["p90"]:.4f} / max={one_thr_stats["max"]:.4f}; '
      f'ratio>=0.5: {ge_05_prop_1hop:.4%}'
    )
    print(
      '(3) two_hop_ratio: '
      f'min={thr_stats["min"]:.4f} / median={thr_stats["median"]:.4f} / mean={thr_stats["mean"]:.4f} / '
      f'p90={thr_stats["p90"]:.4f} / max={thr_stats["max"]:.4f}; '
      f'ratio>=0.9: {ge_09_prop:.4%}'
    )
    print(f'(4) degree quantiles {qs.astype(int).tolist()}: {[round(x, 6) for x in deg_q]}')
    print(f'    one_hop_ratio quantiles {qs.astype(int).tolist()}: {[round(x, 6) for x in thr1_q]}')
    print(f'    two_hop_ratio quantiles {qs.astype(int).tolist()}: {[round(x, 6) for x in thr_q]}')
    print(f'(5) Top-10 degree: {_fmt_list(top_degree)}')
    print(f'    Top-10 one_hop_ratio: {_fmt_list([(a, round(b, 6)) for a, b in top_thr1])}')
    print(f'    Bottom-10 one_hop_ratio: {_fmt_list([(a, round(b, 6)) for a, b in bottom_thr1])}')
    print(f'    Top-10 two_hop_ratio: {_fmt_list([(a, round(b, 6)) for a, b in top_thr])}')
    print(f'    Bottom-10 two_hop_ratio: {_fmt_list([(a, round(b, 6)) for a, b in bottom_thr])}')
    print(f'(6) Spearman(degree, two_hop_ratio)≈Pearson(rank): {corr:.6f}')

  df = pd.DataFrame({
    'node_id': [int(n) for n in nodes],
    'degree': degrees.astype(np.int64),
    'one_hop_count': one_hop_counts.astype(np.int64),
    'one_hop_ratio': one_hop_ratios.astype(np.float64),
    'two_hop_count': two_hop_counts.astype(np.int64),
    'two_hop_ratio': two_hop_ratios.astype(np.float64),
  })

  summary = {
    'graph_name': str(graph_name),
    'N': N,
    'E': E,
    'avg_degree': avg_degree,
    'max_degree': max_degree,
    'one_hop_ratio_min': one_thr_stats['min'],
    'one_hop_ratio_median': one_thr_stats['median'],
    'one_hop_ratio_mean': one_thr_stats['mean'],
    'one_hop_ratio_p90': one_thr_stats['p90'],
    'one_hop_ratio_max': one_thr_stats['max'],
    'two_hop_ratio_min': thr_stats['min'],
    'two_hop_ratio_median': thr_stats['median'],
    'two_hop_ratio_mean': thr_stats['mean'],
    'two_hop_ratio_p90': thr_stats['p90'],
    'two_hop_ratio_max': thr_stats['max'],
    'two_hop_ratio_ge_0.9_proportion': ge_09_prop,
    'degree_two_hop_ratio_spearman': corr,
  }
  return df, summary


def analyze_and_export(
    graphs: Iterable[Any],
    out_xlsx: str = 'coarsegraph_analyze.xlsx',
    threshold: float = 0.0,
) -> str:
  try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
  except Exception as e:
    raise RuntimeError('导出 .xlsx 需要 openpyxl；当前环境不可用。') from e

  out_xlsx = str(out_xlsx)
  out_dir = os.path.dirname(out_xlsx)
  if out_dir:
    os.makedirs(out_dir, exist_ok=True)

  if os.path.exists(out_xlsx):
    wb = load_workbook(out_xlsx)
  else:
    wb = Workbook()
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
      ws0 = wb['Sheet']
      ws0.title = 'SUMMARY'
      ws0.append([
        'graph_name', 'N', 'E', 'avg_degree', 'max_degree',
        'one_hop_ratio_min', 'one_hop_ratio_median', 'one_hop_ratio_mean', 'one_hop_ratio_p90', 'one_hop_ratio_max',
        'two_hop_ratio_min', 'two_hop_ratio_median', 'two_hop_ratio_mean', 'two_hop_ratio_p90', 'two_hop_ratio_max',
        'two_hop_ratio_ge_0.9_proportion', 'degree_two_hop_ratio_spearman'
      ])

  if 'SUMMARY' not in wb.sheetnames:
    ws_sum = wb.create_sheet('SUMMARY', 0)
    header = [
      'graph_name', 'N', 'E', 'avg_degree', 'max_degree',
      'one_hop_ratio_min', 'one_hop_ratio_median', 'one_hop_ratio_mean', 'one_hop_ratio_p90', 'one_hop_ratio_max',
      'two_hop_ratio_min', 'two_hop_ratio_median', 'two_hop_ratio_mean', 'two_hop_ratio_p90', 'two_hop_ratio_max',
      'two_hop_ratio_ge_0.9_proportion', 'degree_two_hop_ratio_spearman'
    ]
    ws_sum.append(header)
  else:
    ws_sum = wb['SUMMARY']
    header = [
      'graph_name', 'N', 'E', 'avg_degree', 'max_degree',
      'one_hop_ratio_min', 'one_hop_ratio_median', 'one_hop_ratio_mean', 'one_hop_ratio_p90', 'one_hop_ratio_max',
      'two_hop_ratio_min', 'two_hop_ratio_median', 'two_hop_ratio_mean', 'two_hop_ratio_p90', 'two_hop_ratio_max',
      'two_hop_ratio_ge_0.9_proportion', 'degree_two_hop_ratio_spearman'
    ]
    for j, v in enumerate(header, start=1):
      ws_sum.cell(row=1, column=j).value = v

  used = set(wb.sheetnames)

  for item in graphs:
    graph_name = None
    hatA_or_G = None
    node_ids = None
    print_summary = True

    if isinstance(item, tuple) and len(item) == 2:
      graph_name, hatA_or_G = item
    elif isinstance(item, dict):
      graph_name = item.get('graph_name', item.get('name', None))
      hatA_or_G = item.get('G', item.get('graph', item.get('hatA', item.get('adj', None))))
      node_ids = item.get('node_ids', None)
      print_summary = bool(item.get('print_summary', True))
    else:
      hatA_or_G = item

    if graph_name is None:
      graph_name = f'graph_{len(wb.sheetnames)}'

    base_sheet = _base_sheet_name(str(graph_name))

    df, summary = analyze_one_graph(
      hatA_or_G=hatA_or_G,
      graph_name=str(graph_name),
      node_ids=node_ids,
      threshold=threshold,
      print_summary=print_summary,
    )

    if base_sheet in wb.sheetnames:
      wb.remove(wb[base_sheet])
      used.discard(base_sheet)
      sheet_name = base_sheet
    else:
      sheet_name = _safe_sheet_name(base_sheet, used=used)
    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
      ws.append(r)

    summary_row = [
      summary['graph_name'], summary['N'], summary['E'], summary['avg_degree'], summary['max_degree'],
      summary['one_hop_ratio_min'], summary['one_hop_ratio_median'], summary['one_hop_ratio_mean'],
      summary['one_hop_ratio_p90'], summary['one_hop_ratio_max'],
      summary['two_hop_ratio_min'], summary['two_hop_ratio_median'], summary['two_hop_ratio_mean'],
      summary['two_hop_ratio_p90'], summary['two_hop_ratio_max'],
      summary['two_hop_ratio_ge_0.9_proportion'], summary['degree_two_hop_ratio_spearman']
    ]
    target_row = None
    for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, min_col=1, max_col=1):
      if row and row[0].value == summary['graph_name']:
        target_row = row[0].row
        break
    if target_row is None:
      ws_sum.append(summary_row)
    else:
      for j, v in enumerate(summary_row, start=1):
        ws_sum.cell(row=target_row, column=j).value = v

  wb.save(out_xlsx)
  return out_xlsx


def default_coarsegraph_analyze_out_xlsx(hparams, data_name=None) -> str:
  if data_name is None:
    data_name = str(getattr(hparams, 'data_name', 'data')).strip() or 'data'
  ts = str(getattr(hparams, 'timestamp', '')).strip() or 'run'
  prefix = f'{data_name}_'
  yml_ts = ts[len(prefix):] if ts.startswith(prefix) and ts[len(prefix):].strip() else ts
  out_dir = '/data/yg/Subgraph-MIL/diffpool2/results'
  try:
    os.makedirs(out_dir, exist_ok=True)
  except Exception:
    out_dir = os.path.join(str(getattr(hparams, 'model_save_path', 'results')), 'coarsegraph_analyze')
    os.makedirs(out_dir, exist_ok=True)
  return os.path.join(out_dir, f'{data_name}_{yml_ts}_coarsegraph_analyze.xlsx')
